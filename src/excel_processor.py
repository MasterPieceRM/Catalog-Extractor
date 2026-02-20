"""
Excel Processing Module
Handles loading, parsing, and batch processing of Excel files for product extraction
"""
import io
import re
import base64
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Generator, Tuple
from dataclasses import dataclass, field
from functools import lru_cache

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


@dataclass
class ExcelImageMeta:
    """Lightweight image metadata — no pixel data stored here"""
    index: int        # Index into the worksheet's _images list
    sheet_name: str   # Which sheet this image belongs to
    row: int          # 0-indexed row where image is anchored
    col: int          # 0-indexed column where image is anchored
    format: str       # Image format (png, jpeg, etc.)
    width: int = 0
    height: int = 0


class ExcelImageStore:
    """
    Lazy image loader — holds the workbook bytes and extracts individual 
    images on demand with an LRU cache. This avoids loading 3500 images 
    into memory at once.
    """

    def __init__(self, file_bytes: bytes):
        self._file_bytes = file_bytes
        # (sheet, index) -> bytes
        self._cache: Dict[Tuple[str, int], bytes] = {}
        self._cache_max = 50  # Keep at most 50 images in memory
        self._cache_order: List[Tuple[str, int]] = []

    def get_image_bytes(self, sheet_name: str, image_index: int) -> Optional[bytes]:
        """Get raw image bytes for a specific image, loading lazily"""
        cache_key = (sheet_name, image_index)

        # Check cache first
        if cache_key in self._cache:
            return self._cache[cache_key]

        # Load from workbook
        try:
            wb = load_workbook(io.BytesIO(self._file_bytes), data_only=True)
            ws = wb[sheet_name]

            if hasattr(ws, '_images') and ws._images and image_index < len(ws._images):
                img = ws._images[image_index]
                data = img._data()
                wb.close()

                # Cache with eviction
                self._cache[cache_key] = data
                self._cache_order.append(cache_key)
                if len(self._cache_order) > self._cache_max:
                    evict_key = self._cache_order.pop(0)
                    self._cache.pop(evict_key, None)

                return data

            wb.close()
        except Exception as e:
            logger.warning(
                f"Failed to lazy-load image {sheet_name}[{image_index}]: {e}")

        return None

    def get_image_base64(self, sheet_name: str, image_index: int) -> Optional[str]:
        """Get base64-encoded image string"""
        data = self.get_image_bytes(sheet_name, image_index)
        if data:
            return base64.b64encode(data).decode('utf-8')
        return None

    def preload_for_rows(self, sheet_name: str, image_metas: List[ExcelImageMeta], row_set: set):
        """
        Batch-preload images for a set of rows into cache.
        Used during export to avoid opening the workbook per-image.
        """
        indices_to_load = []
        for meta in image_metas:
            if meta.sheet_name == sheet_name and meta.row in row_set:
                cache_key = (sheet_name, meta.index)
                if cache_key not in self._cache:
                    indices_to_load.append(meta.index)

        if not indices_to_load:
            return

        try:
            wb = load_workbook(io.BytesIO(self._file_bytes), data_only=True)
            ws = wb[sheet_name]

            if hasattr(ws, '_images') and ws._images:
                for idx in indices_to_load:
                    if idx < len(ws._images):
                        try:
                            data = ws._images[idx]._data()
                            cache_key = (sheet_name, idx)
                            self._cache[cache_key] = data
                            self._cache_order.append(cache_key)
                        except Exception as e:
                            logger.warning(
                                f"Failed to preload image {idx}: {e}")

            wb.close()

            # Evict if over limit (keep generous limit during batch preload)
            while len(self._cache_order) > self._cache_max * 4:
                evict_key = self._cache_order.pop(0)
                self._cache.pop(evict_key, None)

        except Exception as e:
            logger.warning(f"Failed to batch preload images: {e}")


# Keep old ExcelImage for backward compat (e.g. xls format that can't lazy-load)
@dataclass
class ExcelImage:
    """Represents an image from an Excel sheet (legacy, holds bytes in memory)"""
    row: int
    col: int
    image_data: bytes
    format: str
    width: int = 0
    height: int = 0

    def to_base64(self) -> str:
        import base64
        return base64.b64encode(self.image_data).decode('utf-8')


@dataclass
class ExcelSheet:
    """Represents a single sheet in an Excel workbook"""
    name: str
    headers: List[str]
    rows: List[List[Any]]
    total_rows: int
    data_start_excel_row: int = 2  # 1-based Excel row where data starts (after headers)
    image_metas: List[ExcelImageMeta] = field(
        default_factory=list)  # Lightweight metadata only
    images: List[ExcelImage] = field(
        default_factory=list)  # Legacy: only for .xls files
    _row_image_index: Dict[int, ExcelImageMeta] = field(
        default_factory=dict, repr=False)

    def build_row_image_index(self):
        """Build a row → image metadata index for O(1) lookups (exact match)"""
        self._row_image_index = {}
        for meta in self.image_metas:
            # Exact row match only — no fuzzy ±1
            if meta.row not in self._row_image_index:
                self._row_image_index[meta.row] = meta

    def get_image_meta_for_row(self, row_idx: int) -> Optional[ExcelImageMeta]:
        """O(1) image lookup for a row"""
        return self._row_image_index.get(row_idx)

    @property
    def image_count(self) -> int:
        """Total number of images (from either source)"""
        return len(self.image_metas) or len(self.images)

    def get_preview(self, max_rows: int = 20) -> List[List[Any]]:
        """Get a preview of the first N rows"""
        return self.rows[:max_rows]

    def get_rows_as_dicts(self, start_row: int = 0, end_row: int = None) -> List[Dict[str, Any]]:
        """Convert rows to dictionaries using headers as keys"""
        end_row = end_row or len(self.rows)
        result = []
        for row in self.rows[start_row:end_row]:
            row_dict = {}
            for i, header in enumerate(self.headers):
                if i < len(row):
                    row_dict[header] = row[i]
                else:
                    row_dict[header] = None
            result.append(row_dict)
        return result

    def get_image_for_row(self, row_idx: int) -> Optional['ExcelImageMeta']:
        """Get image metadata associated with a row (if any) — O(1) exact match"""
        meta = self._row_image_index.get(row_idx)
        if meta:
            return meta
        # Fallback to legacy images list (exact match)
        for img in self.images:
            if img.row == row_idx:
                return img
        return None

    def get_images_in_range(self, start_row: int, end_row: int) -> List['ExcelImageMeta']:
        """Get all image metas in a row range"""
        if self.image_metas:
            return [m for m in self.image_metas if start_row <= m.row < end_row]
        return [img for img in self.images if start_row <= img.row < end_row]


@dataclass
class ExcelDocument:
    """Represents a loaded Excel workbook"""
    file_path: Path
    file_name: str
    sheets: Dict[str, ExcelSheet] = field(default_factory=dict)
    sheet_names: List[str] = field(default_factory=list)
    image_store: Optional[ExcelImageStore] = field(default=None, repr=False)

    def get_sheet(self, name: str) -> Optional[ExcelSheet]:
        """Get a sheet by name"""
        return self.sheets.get(name)

    def get_first_sheet(self) -> Optional[ExcelSheet]:
        """Get the first sheet"""
        if self.sheet_names:
            return self.sheets.get(self.sheet_names[0])
        return None

    def get_image_bytes(self, sheet_name: str, image_index: int) -> Optional[bytes]:
        """Get image bytes via the lazy store"""
        if self.image_store:
            return self.image_store.get_image_bytes(sheet_name, image_index)
        return None

    def get_image_base64(self, sheet_name: str, image_index: int) -> Optional[str]:
        """Get image as base64 via the lazy store"""
        if self.image_store:
            return self.image_store.get_image_base64(sheet_name, image_index)
        return None


class ExcelProcessor:
    """Handles Excel file loading and batch processing"""

    def __init__(self, default_batch_size: int = 50):
        self.default_batch_size = default_batch_size

    @staticmethod
    def _deduplicate_headers(headers: List[str]) -> List[str]:
        """Append _2, _3, etc. to duplicate column names to prevent crashes."""
        seen: Dict[str, int] = {}
        result = []
        for h in headers:
            if h in seen:
                seen[h] += 1
                result.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 1
                result.append(h)
        return result

    def load_excel(self, file_path: Path, header_row: Optional[int] = None) -> ExcelDocument:
        """Load an Excel file and extract all sheets"""
        logger.info(f"Loading Excel file: {file_path}")

        doc = ExcelDocument(
            file_path=file_path,
            file_name=file_path.name
        )

        try:
            # Load workbook (data_only=True to get values instead of formulas)
            wb = load_workbook(file_path, data_only=True)

            doc.sheet_names = wb.sheetnames

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet = self._process_sheet(ws, sheet_name, header_row=header_row)
                doc.sheets[sheet_name] = sheet
                logger.info(
                    f"Loaded sheet '{sheet_name}': {sheet.total_rows} rows, {len(sheet.headers)} columns")

            wb.close()

        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            raise

        return doc

    def load_excel_from_bytes(self, file_bytes: bytes, file_name: str,
                              header_row: Optional[int] = None) -> ExcelDocument:
        """Load an Excel file from bytes (e.g., from Streamlit upload).

        Args:
            header_row: 1-based Excel row number to use as header.
                        None = auto-detect (first non-empty row).
                        0 = no header (generate Column_1, Column_2, ...).
        """
        logger.info(f"Loading Excel from bytes: {file_name}")

        doc = ExcelDocument(
            file_path=Path(file_name),
            file_name=file_name
        )

        # Determine file format by extension
        is_xls = file_name.lower().endswith(
            '.xls') and not file_name.lower().endswith('.xlsx')

        try:
            if is_xls:
                # Use xlrd for .xls files (Excel 97-2003 format)
                doc = self._load_xls_from_bytes(file_bytes, file_name, header_row=header_row)
            else:
                # Use openpyxl for .xlsx files
                doc = self._load_xlsx_from_bytes(file_bytes, file_name, header_row=header_row)

        except Exception as e:
            # If openpyxl fails, try xlrd as fallback
            if not is_xls and "zip" in str(e).lower():
                logger.warning(
                    f"openpyxl failed, trying xlrd for: {file_name}")
                try:
                    doc = self._load_xls_from_bytes(file_bytes, file_name, header_row=header_row)
                except Exception as e2:
                    logger.error(f"Both openpyxl and xlrd failed: {e2}")
                    raise ValueError(
                        f"Could not load Excel file. Tried both .xlsx and .xls formats. Error: {e}")
            else:
                logger.error(f"Failed to load Excel from bytes: {e}")
                raise

        return doc

    def _load_xlsx_from_bytes(self, file_bytes: bytes, file_name: str,
                              header_row: Optional[int] = None) -> ExcelDocument:
        """Load .xlsx file using openpyxl — images are metadata-only (lazy loaded)"""
        doc = ExcelDocument(
            file_path=Path(file_name),
            file_name=file_name,
            image_store=ExcelImageStore(file_bytes)
        )

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        doc.sheet_names = wb.sheetnames

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet = self._process_sheet(ws, sheet_name, header_row=header_row)

            # Extract image METADATA only — no pixel data loaded
            image_metas = self._extract_image_metadata(ws, sheet_name)
            sheet.image_metas = image_metas
            sheet.build_row_image_index()

            doc.sheets[sheet_name] = sheet
            logger.info(
                f"Loaded sheet '{sheet_name}': {sheet.total_rows} rows, "
                f"{len(sheet.headers)} columns, {len(image_metas)} images (lazy)")

        wb.close()
        return doc

    def _extract_image_metadata(self, ws: Worksheet, sheet_name: str) -> List[ExcelImageMeta]:
        """Extract image metadata only — no bytes loaded. ~100x faster than full extraction."""
        metas = []

        try:
            if hasattr(ws, '_images') and ws._images:
                for idx, img in enumerate(ws._images):
                    try:
                        anchor = img.anchor
                        row = 0
                        col = 0

                        if hasattr(anchor, '_from'):
                            row = anchor._from.row
                            col = anchor._from.col
                        elif hasattr(anchor, 'row'):
                            row = anchor.row
                            col = anchor.col

                        fmt = 'png'
                        if hasattr(img, 'format') and img.format:
                            fmt = img.format.lower()
                        elif hasattr(img, 'ref') and img.ref:
                            if '.jpg' in str(img.ref).lower() or '.jpeg' in str(img.ref).lower():
                                fmt = 'jpeg'

                        width = getattr(img, 'width', 0) or 0
                        height = getattr(img, 'height', 0) or 0

                        metas.append(ExcelImageMeta(
                            index=idx,
                            sheet_name=sheet_name,
                            row=row,
                            col=col,
                            format=fmt,
                            width=int(width) if width else 0,
                            height=int(height) if height else 0
                        ))
                    except Exception as e:
                        logger.warning(
                            f"Failed to read image metadata {idx}: {e}")
                        continue
        except Exception as e:
            logger.warning(f"Failed to extract image metadata: {e}")

        logger.info(f"Indexed {len(metas)} image positions from sheet")
        return metas

    def _load_xls_from_bytes(self, file_bytes: bytes, file_name: str,
                             header_row: Optional[int] = None) -> ExcelDocument:
        """Load .xls file using xlrd"""
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "xlrd is required for .xls files. Install with: pip install xlrd")

        doc = ExcelDocument(
            file_path=Path(file_name),
            file_name=file_name
        )

        wb = xlrd.open_workbook(file_contents=file_bytes)

        doc.sheet_names = wb.sheet_names()

        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            sheet = self._process_xls_sheet(ws, sheet_name, header_row=header_row)
            doc.sheets[sheet_name] = sheet
            logger.info(
                f"Loaded sheet '{sheet_name}': {sheet.total_rows} rows, {len(sheet.headers)} columns")

        return doc

    def _process_xls_sheet(self, ws, sheet_name: str,
                           header_row: Optional[int] = None) -> ExcelSheet:
        """Process an xlrd worksheet and extract headers and rows.

        Args:
            header_row: 1-based row number for headers.
                        None = auto-detect, 0 = no header.
        """
        rows = []
        headers = []
        num_cols = ws.ncols

        if header_row == 0:
            # No header mode — generate Column_1, Column_2, ...
            headers = [f"Column_{i+1}" for i in range(num_cols)]
            # Keep ALL rows (including empty ones) to preserve structure
            for row_idx in range(ws.nrows):
                row_values = [self._clean_cell_value(ws.cell_value(row_idx, col_idx))
                              for col_idx in range(num_cols)]
                rows.append(row_values)
            data_start = 1  # All rows are data, starting at Excel row 1
        elif header_row is not None and header_row > 0:
            # Specific header row (1-based)
            hr_idx = header_row - 1  # Convert to 0-based
            if hr_idx < ws.nrows:
                row_values = [self._clean_cell_value(ws.cell_value(hr_idx, col_idx))
                              for col_idx in range(num_cols)]
                headers = [str(v) if v else f"Column_{i+1}" for i, v in enumerate(row_values)]
            else:
                headers = [f"Column_{i+1}" for i in range(num_cols)]
            # Keep ALL data rows (including empty) except the header row
            for row_idx in range(ws.nrows):
                if row_idx == hr_idx:
                    continue  # Skip the header row
                row_values = [self._clean_cell_value(ws.cell_value(row_idx, col_idx))
                              for col_idx in range(num_cols)]
                rows.append(row_values)
            data_start = header_row + 1
        else:
            # Auto-detect: first non-empty row = header, then keep ALL rows after
            # If the candidate header row has gaps (empty cells), fall back to no-header mode
            header_found = False
            for row_idx in range(ws.nrows):
                row_values = [self._clean_cell_value(ws.cell_value(row_idx, col_idx))
                              for col_idx in range(num_cols)]
                if not header_found:
                    # Skip leading empty rows
                    if all(v == '' or v is None for v in row_values):
                        continue
                    # Check for gaps between first and last column with data
                    # (trailing empty columns are OK — they're just extra columns)
                    last_data_col = max((i for i, v in enumerate(row_values) if v != '' and v is not None), default=-1)
                    if last_data_col >= 0:
                        core_values = row_values[:last_data_col + 1]
                        has_gap = any(v == '' or v is None for v in core_values)
                    else:
                        has_gap = True
                    if has_gap:
                        # Fall back to no-header mode
                        headers = [f"Column_{i+1}" for i in range(num_cols)]
                        header_found = True
                        rows.append(row_values)  # Keep this row as data
                        data_start = 1
                        continue
                    headers = [
                        str(v) if v else f"Column_{i+1}" for i, v in enumerate(row_values)]
                    header_found = True
                    data_start = 2  # xls: header is this row, data starts next
                else:
                    rows.append(row_values)
            if not header_found:
                data_start = 1

        headers = self._deduplicate_headers(headers)

        return ExcelSheet(
            name=sheet_name,
            headers=headers,
            rows=rows,
            total_rows=len(rows),
            data_start_excel_row=data_start,
        )

    def _process_sheet(self, ws: Worksheet, sheet_name: str,
                       header_row: Optional[int] = None) -> ExcelSheet:
        """Process a worksheet and extract headers and rows.

        Args:
            header_row: 1-based Excel row number for headers.
                        None = auto-detect (first non-empty row).
                        0 = no header (generate Column_1, Column_2, ...).
        """
        rows = []
        headers = []
        header_excel_row = 0  # Track which Excel row is the header

        # Read all rows from the worksheet
        all_rows = list(ws.iter_rows(values_only=True))
        num_cols = ws.max_column or 0

        if header_row == 0:
            # No header mode — generate Column_1, Column_2, ...
            if num_cols == 0 and all_rows:
                num_cols = len(all_rows[0])
            headers = [f"Column_{i+1}" for i in range(num_cols)]
            # Keep ALL rows (including empty) to preserve structure
            for row in all_rows:
                rows.append([self._clean_cell_value(cell) for cell in row])
            data_start_excel_row = 1  # All rows are data
        elif header_row is not None and header_row > 0:
            # Specific header row (1-based, so index = header_row - 1)
            hr_idx = header_row - 1
            if hr_idx < len(all_rows):
                header_values = [self._clean_cell_value(cell) for cell in all_rows[hr_idx]]
                headers = [str(v) if v else f"Column_{i+1}" for i, v in enumerate(header_values)]
            else:
                if num_cols == 0 and all_rows:
                    num_cols = len(all_rows[0])
                headers = [f"Column_{i+1}" for i in range(num_cols)]
            # Keep ALL data rows (including empty) except header row
            for row_idx, row in enumerate(all_rows):
                if row_idx == hr_idx:
                    continue
                rows.append([self._clean_cell_value(cell) for cell in row])
            header_excel_row = hr_idx
            data_start_excel_row = header_row + 1
        else:
            # Auto-detect: first non-empty row = header, then keep ALL rows after
            # If the candidate header row has gaps (empty cells), fall back to no-header mode
            header_found = False
            no_header_fallback = False
            for row_idx, row in enumerate(all_rows):
                if not header_found:
                    # Skip leading empty rows
                    if all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    row_values = [self._clean_cell_value(cell) for cell in row]
                    # Check for gaps between first and last column with data
                    # (trailing empty columns are OK — they're just extra columns)
                    last_data_col = max((i for i, v in enumerate(row_values) if v != '' and v is not None), default=-1)
                    if last_data_col >= 0:
                        core_values = row_values[:last_data_col + 1]
                        has_gap = any(v == '' or v is None for v in core_values)
                    else:
                        has_gap = True
                    if has_gap:
                        # Fall back to no-header mode
                        headers = [f"Column_{i+1}" for i in range(num_cols or len(row))]
                        header_found = True
                        no_header_fallback = True
                        rows.append(row_values)  # Keep this row as data
                        continue
                    headers = [
                        str(v) if v else f"Column_{i+1}" for i, v in enumerate(row_values)]
                    header_excel_row = row_idx  # 0-based
                    header_found = True
                else:
                    rows.append([self._clean_cell_value(cell) for cell in row])
            if no_header_fallback:
                data_start_excel_row = 1  # All rows are data
            else:
                data_start_excel_row = header_excel_row + 2  # 1-based, row after header

        headers = self._deduplicate_headers(headers)

        return ExcelSheet(
            name=sheet_name,
            headers=headers,
            rows=rows,
            total_rows=len(rows),
            data_start_excel_row=data_start_excel_row,
        )

    def _clean_cell_value(self, value: Any) -> Any:
        """Clean and normalize a cell value"""
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            # Remove trailing .0 for whole numbers
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return value
        return str(value).strip()

    def get_batches(
        self,
        sheet: ExcelSheet,
        batch_size: int = None,
        start_row: int = 0,
        end_row: int = None
    ) -> Generator[Tuple[int, int, List[Dict[str, Any]]], None, None]:
        """
        Generate batches of rows from a sheet.

        Yields tuples of (batch_start, batch_end, rows_as_dicts)
        """
        batch_size = batch_size or self.default_batch_size
        end_row = end_row or sheet.total_rows

        current_start = start_row
        while current_start < end_row:
            current_end = min(current_start + batch_size, end_row)
            rows = sheet.get_rows_as_dicts(current_start, current_end)
            yield (current_start, current_end, rows)
            current_start = current_end

    def rows_to_text(self, rows: List[Dict[str, Any]], headers: List[str]) -> str:
        """
        Convert rows to a text format suitable for LLM processing.
        Uses a structured format that preserves column context.
        """
        lines = []

        # Add header context
        lines.append(f"=== SPREADSHEET DATA ({len(rows)} rows) ===")
        lines.append(f"Columns: {', '.join(headers)}")
        lines.append("")

        for row_idx, row in enumerate(rows):
            lines.append(f"--- Row {row_idx + 1} ---")
            for header in headers:
                value = row.get(header, "")
                if value and str(value).strip():
                    lines.append(f"  {header}: {value}")
            lines.append("")

        return "\n".join(lines)

    def get_total_batches(
        self,
        sheet: ExcelSheet,
        batch_size: int = None,
        start_row: int = 0,
        end_row: int = None
    ) -> int:
        """Calculate how many batches will be generated"""
        batch_size = batch_size or self.default_batch_size
        end_row = end_row or sheet.total_rows
        total_rows = end_row - start_row
        return (total_rows + batch_size - 1) // batch_size  # Ceiling division

    def get_sample_rows(self, sheet: ExcelSheet, num_samples: int = 5, start_row: int = 0) -> List[List[Any]]:
        """Get a representative sample of rows for pattern learning"""
        total = len(sheet.rows)
        if total == 0:
            return []

        # Take rows spread across the data for a representative sample
        if total <= num_samples:
            return sheet.rows[start_row:]

        # Take first few and some from middle to get diversity
        indices = [start_row]
        step = max(1, (total - start_row) // num_samples)
        for i in range(1, num_samples):
            idx = start_row + i * step
            if idx < total:
                indices.append(idx)

        return [sheet.rows[i] for i in indices if i < total]

    def apply_pattern(
        self,
        sheet: ExcelSheet,
        mapping: Dict[str, Any],
        start_row: int = 0,
        end_row: int = None,
    ) -> List[Dict[str, Any]]:
        """
        Apply a learned column mapping to all rows, producing product dicts.
        No LLM calls — pure Python transformation.

        Args:
            sheet: The Excel sheet to process
            mapping: The mapping dict from learn_excel_pattern()
            start_row: Start row index (0-based into sheet.rows)
            end_row: End row index (exclusive, 0-based into sheet.rows)

        Returns:
            List of product dicts ready to be converted to Product objects.
            Each dict has _source_row/_source_rows as Excel-absolute 1-based row numbers.
        """
        end_row = end_row or sheet.total_rows
        column_mappings = mapping.get("column_mappings", {})
        row_grouping = mapping.get("row_grouping", "one_per_row")
        group_key_col = mapping.get("group_key_column")

        # Build header index for fast lookup
        header_idx = {h: i for i, h in enumerate(sheet.headers)}

        rows_slice = sheet.rows[start_row:end_row]

        # Excel-absolute offset: data_start_excel_row is the 1-based Excel row of the first data row
        excel_offset = sheet.data_start_excel_row + start_row

        if row_grouping == "one_per_row":
            return self._apply_one_per_row(rows_slice, column_mappings, header_idx, excel_offset)
        else:
            return self._apply_multi_row(rows_slice, column_mappings, header_idx, group_key_col, excel_offset)

    def _apply_one_per_row(
        self,
        rows: List[List[Any]],
        column_mappings: Dict[str, Any],
        header_idx: Dict[str, int],
        start_row_offset: int
    ) -> List[Dict[str, Any]]:
        """Apply mapping where each row is one product"""
        products = []

        for row_idx, row in enumerate(rows):
            product_dict = {}
            has_any_value = False

            for field_name, field_mapping in column_mappings.items():
                value = self._extract_field_value(
                    row, field_mapping, header_idx)
                if value is not None and str(value).strip():
                    has_any_value = True
                product_dict[field_name] = value

            # Skip completely empty rows
            if not has_any_value:
                continue

            # Ensure there's at least a name
            if not product_dict.get("name"):
                # Try to use the first non-empty value as name
                for v in row:
                    if v and str(v).strip():
                        product_dict["name"] = str(v)
                        break

            product_dict["_source_row"] = start_row_offset + \
                row_idx  # Excel-absolute row number
            products.append(product_dict)

        return products

    def _apply_multi_row(
        self,
        rows: List[List[Any]],
        column_mappings: Dict[str, Any],
        header_idx: Dict[str, int],
        group_key_col: Optional[str],
        start_row_offset: int
    ) -> List[Dict[str, Any]]:
        """Apply mapping where multiple rows form one product"""
        groups: Dict[str, Dict[str, Any]] = {}
        group_rows: Dict[str, List[int]] = {}
        group_order = []

        key_idx = header_idx.get(group_key_col) if group_key_col else None

        for row_idx, row in enumerate(rows):
            # Determine group key
            if key_idx is not None and key_idx < len(row):
                key = str(row[key_idx]).strip()
                if not key:
                    # Empty key — append to previous group if exists
                    key = group_order[-1] if group_order else f"_row_{row_idx}"
            else:
                key = f"_row_{row_idx}"

            if key not in groups:
                groups[key] = {}
                group_rows[key] = []
                group_order.append(key)

            group_rows[key].append(start_row_offset + row_idx)

            # Merge values into group
            for field_name, field_mapping in column_mappings.items():
                value = self._extract_field_value(
                    row, field_mapping, header_idx)
                if value is not None and str(value).strip():
                    existing = groups[key].get(field_name)
                    if not existing or not str(existing).strip():
                        groups[key][field_name] = value
                    elif field_name in ("description", "features"):
                        # Concatenate for text fields
                        groups[key][field_name] = f"{existing}; {value}"

        products = []
        for key in group_order:
            product_dict = groups[key]
            product_dict["_source_rows"] = group_rows[key]
            if product_dict.get("name") or any(str(v).strip() for v in product_dict.values() if v):
                products.append(product_dict)

        return products

    def _extract_field_value(
        self,
        row: List[Any],
        field_mapping: Dict[str, Any],
        header_idx: Dict[str, int]
    ) -> Any:
        """Extract a single field value from a row using the mapping rules"""
        # Check for fixed value first
        fixed = field_mapping.get("fixed_value")
        if fixed is not None:
            return fixed

        source_columns = field_mapping.get("source_columns", [])
        transform = field_mapping.get("transform", "none")

        if not source_columns:
            return None

        # Gather raw values from source columns
        raw_values = []
        for col_name in source_columns:
            idx = header_idx.get(col_name)
            if idx is not None and idx < len(row):
                raw_values.append(row[idx])
            else:
                raw_values.append(None)

        # Apply transform
        if transform == "extract_number":
            return self._transform_extract_number(raw_values)
        elif transform == "concatenate":
            return self._transform_concatenate(raw_values)
        elif transform == "first_non_empty":
            return self._transform_first_non_empty(raw_values)
        else:  # "none"
            return self._transform_first_non_empty(raw_values)

    def _transform_extract_number(self, values: List[Any]) -> Optional[float]:
        """Extract numeric value from text (removes currency symbols, etc.)"""
        for v in values:
            if v is None:
                continue
            if isinstance(v, (int, float)):
                return float(v)
            text = str(v).strip()
            if not text:
                continue
            # Remove currency symbols and whitespace
            cleaned = re.sub(r'[€$£¥CHF\s]', '', text)
            # Handle European format: 1.234,56 -> 1234.56
            if ',' in cleaned and '.' in cleaned:
                if cleaned.index('.') < cleaned.index(','):
                    cleaned = cleaned.replace('.', '').replace(',', '.')
                else:
                    cleaned = cleaned.replace(',', '')
            elif ',' in cleaned:
                cleaned = cleaned.replace(',', '.')
            try:
                return float(cleaned)
            except ValueError:
                # Try to find any number in the string
                match = re.search(r'[\d]+[.,]?\d*', cleaned)
                if match:
                    try:
                        return float(match.group().replace(',', '.'))
                    except ValueError:
                        pass
        return None

    def _transform_concatenate(self, values: List[Any]) -> str:
        """Concatenate non-empty values"""
        parts = [str(v).strip()
                 for v in values if v is not None and str(v).strip()]
        return " ".join(parts) if parts else ""

    def _transform_first_non_empty(self, values: List[Any]) -> Any:
        """Return the first non-empty value"""
        for v in values:
            if v is not None and str(v).strip():
                return v
        return None
