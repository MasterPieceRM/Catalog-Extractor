"""
Excel Export Module
Exports products to Excel with embedded images using openpyxl
"""
import io
import base64
import logging
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports products to Excel with embedded images"""

    # Image size presets (width, height) in pixels
    IMAGE_SIZES = {
        'small': (60, 60),
        'medium': (100, 100),
        'large': (150, 150)
    }

    def __init__(self):
        self.header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export_products_to_excel(
        self,
        products: List,
        schema_fields: List[Dict[str, Any]],
        include_images: bool = True,
        image_size: str = 'medium',
        include_page_info: bool = True,
        remove_bg: bool = False,
        include_row_info: bool = False,
        computed_fields: List[Dict[str, str]] = None
    ) -> bytes:
        """
        Export products to Excel with optional embedded images.
        
        Args:
            computed_fields: List of dicts with 'name' and 'formula' keys.
                             Example: [{'name': 'total', 'formula': '{price}*2'}]
        """
        # Note: rembg is imported inside the loop with warning suppression if remove_bg is True

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        img_width, img_height = self.IMAGE_SIZES.get(
            image_size, self.IMAGE_SIZES['medium'])

        # Build headers
        headers = []
        if include_images:
            headers.append("Image")

        for field in schema_fields:
            headers.append(field['name'].replace('_', ' ').title())

        # Computed fields headers
        if computed_fields:
            for field in computed_fields:
                headers.append(field['name'].replace('_', ' ').title())

        if include_page_info:
            headers.append("Page")

        if include_row_info:
            headers.append("Source Rows")

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # Set column widths
        img_col_width = img_width / 7 + 2 if include_images else 0  # Approximate Excel units
        for col, header in enumerate(headers, 1):
            if include_images and col == 1:
                ws.column_dimensions[get_column_letter(
                    col)].width = img_col_width
            else:
                ws.column_dimensions[get_column_letter(col)].width = 25

        # Write product data
        current_row = 2
        for product in products:
            col_offset = 1

            # Add image if available
            if include_images:
                row_height = img_height * 0.75  # Approximate Excel row height units
                ws.row_dimensions[current_row].height = row_height

                product_image = self._get_product_image(product)

                # Apply background removal if requested
                if product_image and remove_bg:
                    try:
                        # Suppress ONNX Runtime and Numba warnings
                        import warnings
                        import os
                        os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
                        with warnings.catch_warnings():
                            warnings.simplefilter("ignore")
                            import rembg
                            product_image = rembg.remove(product_image)
                    except Exception as e:
                        logger.error(
                            f"Failed to remove background in export: {e}")

                if product_image:
                    try:
                        # Resize image
                        resized = self._resize_image(
                            product_image, img_width, img_height)

                        # Save to bytes
                        img_bytes = io.BytesIO()
                        resized.save(img_bytes, format='PNG')
                        img_bytes.seek(0)

                        # Create Excel image
                        xl_img = XLImage(img_bytes)
                        xl_img.width = resized.width
                        xl_img.height = resized.height

                        # Position image in cell
                        cell_ref = f"A{current_row}"
                        ws.add_image(xl_img, cell_ref)

                        # Center align image cell (affects text if any fallback)
                        cell = ws.cell(row=current_row, column=1)
                        cell.alignment = Alignment(
                            horizontal='center', vertical='center')

                    except Exception as e:
                        logger.error(f"Failed to embed image: {e}")
                        ws.cell(row=current_row, column=1,
                                value="[Image Error]")
                else:
                    ws.cell(row=current_row, column=1, value="")

                col_offset = 2

            # Add field values
            for field_idx, field in enumerate(schema_fields):
                field_name = field['name']
                col = col_offset + field_idx

                # Get value from product
                value = self._get_field_value(product, field_name)

                cell = ws.cell(row=current_row, column=col, value=value)
                # Centered alignment for ALL cells
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.border

            # Add page number
            if include_page_info:
                page_col = col_offset + len(schema_fields)
                cell = ws.cell(row=current_row, column=page_col,
                               value=product.page_number + 1)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.border = self.border

            # Add source row info
            if include_row_info:
                row_info_col = col_offset + \
                    len(schema_fields) + (1 if include_page_info else 0)
                source_rows = getattr(product, 'source_rows', [])
                row_info_value = ', '.join(
                    map(str, source_rows)) if source_rows else ''
                cell = ws.cell(row=current_row,
                               column=row_info_col, value=row_info_value)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.border = self.border

            current_row += 1

        # === Apply formulas for computed fields ===
        if computed_fields:
            # Build field_name -> column_letter mapping for ALL fields (schema + computed)
            field_col_map = {}
            col_offset = 2 if include_images else 1
            
            # Map schema fields
            for field_idx, field in enumerate(schema_fields):
                col_num = col_offset + field_idx
                field_col_map[field['name']] = get_column_letter(col_num)
            
            # Map computed fields (they come after schema fields)
            computed_start_col = col_offset + len(schema_fields)
            for idx, field in enumerate(computed_fields):
                col_num = computed_start_col + idx
                field_col_map[field['name']] = get_column_letter(col_num)

            # Apply each formula to every product row
            from openpyxl.utils import column_index_from_string
            data_start_row = 2
            data_end_row = data_start_row + len(products) - 1

            for field in computed_fields:
                name = field['name']
                formula_template = field['formula']
                
                if not formula_template:
                    continue

                target_col = field_col_map.get(name)
                if not target_col:
                    continue
                    
                target_col_num = column_index_from_string(target_col)

                for row_num in range(data_start_row, data_end_row + 1):
                    # Replace {field_name} references with cell references
                    excel_formula = formula_template
                    for ref_name, ref_col in field_col_map.items():
                        # Use word boundary check if needed, but simple replace usually works for field names
                        # Better: sort by length desc to avoid prefix matching issues
                        pass
                    
                    # Sort keys by length descending to avoid partial replacements (e.g. {price} vs {price_total})
                    sorted_refs = sorted(field_col_map.items(), key=lambda x: len(x[0]), reverse=True)
                    
                    for ref_name, ref_col in sorted_refs:
                        excel_formula = excel_formula.replace(
                            '{' + ref_name + '}', f'{ref_col}{row_num}')

                    # Write as Excel formula
                    cell = ws.cell(row=row_num, column=target_col_num)
                    cell.value = f'={excel_formula}'
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center', wrap_text=True)
                    cell.border = self.border

            # Mark computed column headers with a visual indicator
            formula_header_fill = PatternFill(
                start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            formula_header_font = Font(bold=True, color="000000", size=11)
            
            for field in computed_fields:
                name = field['name']
                if name in field_col_map:
                    col_letter = field_col_map[name]
                    col_num = column_index_from_string(col_letter)
                    header_cell = ws.cell(row=1, column=col_num)
                    header_cell.fill = formula_header_fill
                    header_cell.font = formula_header_font

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _get_product_image(self, product) -> Optional[Image.Image]:
        """Get the first available image from a product"""
        # First check for excel_image (from Excel extraction)
        excel_image = getattr(product, 'excel_image', None)
        if excel_image:
            try:
                image_bytes = base64.b64decode(excel_image)
                return Image.open(io.BytesIO(image_bytes))
            except Exception as e:
                logger.error(f"Failed to decode excel_image: {e}")

        # Fall back to images list (from PDF extraction)
        if not product.images:
            return None

        for img in product.images:
            if img.image_data:
                try:
                    image_bytes = base64.b64decode(img.image_data)
                    return Image.open(io.BytesIO(image_bytes))
                except Exception as e:
                    logger.error(f"Failed to decode product image: {e}")
                    continue

        return None

    def _get_field_value(self, product, field_name: str) -> Any:
        """Get a field value from a product"""
        # Check direct attributes first
        if hasattr(product, field_name):
            value = getattr(product, field_name)
        else:
            value = product.raw_attributes.get(field_name, '')

        # Handle special types
        if field_name == 'price' and value:
            if hasattr(value, 'amount'):
                return value.amount
            return str(value)

        if isinstance(value, list):
            return ', '.join(str(v) for v in value)

        if isinstance(value, dict):
            return str(value)

        return value if value is not None else ''

    def _resize_image(
        self,
        image: Image.Image,
        max_width: int,
        max_height: int
    ) -> Image.Image:
        """Resize image maintaining aspect ratio"""
        # Convert to RGB if necessary (for PNG with transparency)
        if image.mode == 'P':
            image = image.convert('RGBA')

        if image.mode == 'RGBA':
            # Keep transparency
            pass
        elif image.mode != 'RGB':
            image = image.convert('RGB')

        ratio = min(max_width / image.width, max_height / image.height)
        if ratio < 1:
            new_size = (int(image.width * ratio), int(image.height * ratio))
            return image.resize(new_size, Image.Resampling.LANCZOS)
        return image


def create_excel_export(
    products: List,
    schema_fields: List[Dict[str, Any]],
    computed_fields: List[Dict[str, str]] = None,
    include_images: bool = True,
    image_size: str = 'medium'
) -> bytes:
    """
    Convenience function to create an Excel export.

    Args:
        products: List of Product objects
        schema_fields: Schema field definitions
        computed_fields: List of computed field definitions
        include_images: Whether to include images
        image_size: Image size preset

    Returns:
        Excel file as bytes
    """
    exporter = ExcelExporter()
    return exporter.export_products_to_excel(
        products=products,
        schema_fields=schema_fields,
        computed_fields=computed_fields,
        include_images=include_images,
        image_size=image_size
    )
